import json
from openpyxl import Workbook
import re
def main():
    with open(f'sample.json') as file:
        datas = json.load(file)
    load_excel(datas)
def load_excel(json_file_name: str, excel_file_name: str,datas):
    workbook = Workbook()
    workbook_active = workbook.active

    categories = ['Company', 'URL', 'Role', 'Location', 'Date Posted']

    for col_num, category in enumerate(categories, start=1):
        workbook_active.cell(row=1, column=col_num, value=category)
    for index, data in enumerate(datas):
        workbook_active.cell(row=index+2, column=1, value=' '.join(datas[index]['company']).title())
        workbook_active.cell(row=index+2, column=2, value=' '.join(datas[index]['url']))
        workbook_active.cell(row=index+2, column=3, value=' '.join(datas[index]['role']).title())
        workbook_active.cell(row=index+2, column=4, value=' '.join(datas[index]['location']).upper())
        workbook_active.cell(row=index+2, column=5, value=' '.join(datas[index]['date posted']).title())
    workbook.save(f"{excel_file_name}.xlsx")
if __name__ == "__main__":
    main()