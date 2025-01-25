import json
from openpyxl import Workbook
import re
def main():

    workbook = Workbook()
    workbook_active = workbook.active

    categories = ['Company', 'URL', 'Role', 'Location', 'Date Posted']

    with open('swe-data.json') as file:
        datas = json.load(file)
    for col_num, category in enumerate(categories, start=1):
        workbook_active.cell(row=1, column=col_num, value=category)
    for index, data in enumerate(datas):
        workbook_active.cell(row=index+2, column=1, value=' '.join(datas[index]['Company']))
        workbook_active.cell(row=index+2, column=2, value=' '.join(datas[index]['URL']))
        workbook_active.cell(row=index+2, column=3, value=' '.join(datas[index]['Role']))
        workbook_active.cell(row=index+2, column=4, value=' '.join(datas[index]['Location']))
        workbook_active.cell(row=index+2, column=5, value=' '.join(datas[index]['Date Posted']))
    workbook.save("swe-list.xlsx")
    test_val = 1

if __name__ == "__main__":
    main()